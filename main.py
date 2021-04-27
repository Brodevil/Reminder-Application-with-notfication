import datetime
import time 
from plyer import notification
import openpyxl


def readExel(path):
    """Read the Exel file using openpyxl and return the 2d list as 1st will the time and 2nd will the purpose """

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    records = list()

    # Loop will print all columns name
    for j in range(1, sheet_obj.max_row + 1):
        for i in range(1, max_col + 1):
            a = sheet_obj.cell(row=j, column=i)
            records.append(a.value)
    records = [[records[i], records[i + 1]] for i in range(0, len(records), 2)]
    return records


def reminder(data):
    """Funtion compare the time and just give notification if the saved time is same as the present time"""\

    while True:
        for excelTime, pourpose in data:
            currentTime = datetime.time(int(datetime.datetime.now().strftime("%H")), int(datetime.datetime.now().strftime("%M")))
            if excelTime == currentTime:
                print(pourpose)
                notification.notify(
                    title=pourpose,
                    message="Sir, You should do your work at the time and plz don't waste the time sir as, A time is a important role in the programmer life. Keep practicing.",
                    app_icon=r"Notification.ico",
                    timeout=12
                )
                time.sleep(62)
                

# def API(url):
    # url = r"https://mashape-community-urban-dictionary.p.rapidapi.com/define"
    # import requests
    # import json
    # querystring = {"term":"wat"}
    # headers = {
    #     'x-rapidapi-key': "03aa284a3fmshda0df2e709f3afap11bcf0jsnf13dd3d99fbb",
    #     'x-rapidapi-host': "mashape-community-urban-dictionary.p.rapidapi.com"
    #     }

    # response = requests.request("GET", url, headers=headers, params=querystring)
    # import pprint
    # pprint.pprint(response.text)


if __name__ == "__main__":
    # Giving the excel file's path as a parameters
    exel = readExel("M:\ADMIN\Critical Data\VS-Code\Reminder Application with Notification\Remind_data.xlsx")
    reminder(exel)
